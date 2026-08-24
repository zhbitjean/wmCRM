"""Import the Bing jobsite measurement tracker into WM CRM without duplicating sites."""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database import SessionLocal
from app.models import Contact, Project, ProjectContact, Property, VerificationStatus


SOURCE = "Bing_Jobsite_meaurements_tracking_list.xlsx / Sheet1"
SOURCE_MARKER = f"Imported from {SOURCE}"
AUDIT_USER = "zhbitjean@gmail.com"


def text(value) -> str:
    return str(value).strip() if value is not None else ""


def key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return text(value)


def parse_address(raw: str, boro_city: str) -> dict[str, str | None]:
    cleaned = re.sub(r"\s+", " ", raw.strip().replace(", USA", ""))
    parts = [part.strip() for part in cleaned.split(",") if part.strip()]
    street = parts[0]
    remainder = ", ".join(parts[1:])
    match = re.search(r"(?P<city>[^,]+?),?\s+NY\s*,?\s*(?P<zip>\d{5})?$", remainder, re.I)
    city = (match.group("city").strip() if match else text(boro_city)) or "New York"
    zip_code = (match.group("zip") or "") if match else ""
    city = {"Brookly": "Brooklyn", "Manhattan": "New York"}.get(city, city)
    locality = text(boro_city)
    if locality in {"Brooklyn", "Brookly"} or city == "Brooklyn":
        borough = "Brooklyn"
    elif locality in {"New York", "Manhattan"} or city in {"New York", "Manhattan"}:
        borough = "Manhattan"
    elif locality == "Staten Island" or city == "Staten Island":
        borough = "Staten Island"
    elif locality in {"Long Island City", "Astoria", "Jamaica", "Rego Park", "Flushing", "Far Rockaway"}:
        borough = "Queens"
    else:
        borough = None
    return {"street": street, "city": city, "state": "NY", "zip": zip_code, "borough": borough}


def status_for(rows: list[dict]) -> str:
    latest = sorted(rows, key=lambda row: row["date"], reverse=True)[0]
    combined = f'{latest["job_status"]} {latest["drawing_status"]}'.lower()
    if "cancel" in combined:
        return "Cancelled"
    if any(word in combined for word in ("done", "submitted", "approved")):
        return "Completed"
    return "Active"


def project_type_for(rows: list[dict]) -> str | None:
    types = list(dict.fromkeys(row["job_type"] for row in rows if row["job_type"]))
    return "; ".join(types)[:100] or None


def tracker_notes(rows: list[dict]) -> str:
    lines = [SOURCE_MARKER]
    for row in sorted(rows, key=lambda item: item["date"]):
        details = []
        for label, field in (
            ("job type", "job_type"), ("job status", "job_status"),
            ("drawing", "drawing_status"), ("client", "client"),
            ("scope", "scope"), ("comments", "comments"),
        ):
            if row[field]:
                details.append(f'{label}: {row[field]}')
        if details:
            lines.append(f'{row["date"] or "date not recorded"}: ' + "; ".join(details))
        else:
            lines.append(f'{row["date"] or "date not recorded"}: measurement recorded')
    return "\n".join(lines)


def merge_import_notes(existing: str | None, imported: str) -> str:
    if not existing:
        return imported
    if SOURCE_MARKER in existing:
        return existing.split(SOURCE_MARKER, 1)[0].rstrip() + "\n\n" + imported
    return existing.rstrip() + "\n\n" + imported


def contact_candidates(db) -> dict[str, Contact]:
    result = {}
    for contact in db.scalars(select(Contact)).all():
        for value in (contact.display_name, contact.first_name, contact.nickname):
            if value:
                result.setdefault(key(value), contact)
    # Known tracker spelling for Joseph Aghelian's nickname Gongji.
    if "gongji" in result:
        result["gojing"] = result["gongji"]
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--commit", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True, read_only=True)
    sheet = workbook["Sheet1"]
    headers = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    grouped: dict[str, list[dict]] = defaultdict(list)
    addresses: dict[str, dict] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        raw_address = text(row.get("Jobsite"))
        if not raw_address:
            continue
        address = parse_address(raw_address, text(row.get("Boro/City")))
        address_key = key(address["street"])
        grouped[address_key].append({
            "date": format_date(row.get("Measurement Date")),
            "job_type": text(row.get("Job Type")),
            "job_status": text(row.get("Job Status")),
            "drawing_status": text(row.get("Drawing Status")),
            "client": text(row.get("Client")),
            "scope": text(row.get("Scope of Work")),
            "comments": text(row.get("Comments")),
        })
        addresses.setdefault(address_key, address)

    db = SessionLocal()
    try:
        existing_properties = db.scalars(select(Property)).all()
        by_street = {key(prop.street_address.split(",")[0]): prop for prop in existing_properties}
        contacts = contact_candidates(db)
        created_contacts = created_properties = created_projects = updated_projects = links = 0

        for address_key, rows in grouped.items():
            address = addresses[address_key]
            prop = by_street.get(address_key)
            if not prop:
                prop = Property(
                    street_address=address["street"], city=address["city"], borough=address["borough"],
                    state=address["state"], zip_code=address["zip"], notes=SOURCE_MARKER,
                    created_by=AUDIT_USER, updated_by=AUDIT_USER,
                )
                db.add(prop)
                db.flush()
                by_street[address_key] = prop
                created_properties += 1

            project = db.scalar(select(Project).where(Project.property_id == prop.id).order_by(Project.id))
            notes = tracker_notes(rows)
            if not project:
                project = Project(
                    project_name=f'{address["street"]} Measurement', property=prop,
                    project_type=project_type_for(rows), status=status_for(rows),
                    internal_notes=notes, created_by=AUDIT_USER, updated_by=AUDIT_USER,
                )
                db.add(project)
                db.flush()
                created_projects += 1
            else:
                project.internal_notes = merge_import_notes(project.internal_notes, notes)
                project.project_type = project.project_type or project_type_for(rows)
                if project.status == "Active":
                    project.status = status_for(rows)
                project.updated_by = AUDIT_USER
                updated_projects += 1

            for name in dict.fromkeys(row["client"] for row in rows if row["client"]):
                contact = contacts.get(key(name))
                if not contact:
                    name_parts = name.split(None, 1)
                    contact = Contact(
                        first_name=name_parts[0], last_name=name_parts[1] if len(name_parts) > 1 else "",
                        display_name=name, role="Client", notes=f"Name imported from {SOURCE}; contact details not provided.",
                        verification_status=VerificationStatus.PENDING,
                        created_by=AUDIT_USER, updated_by=AUDIT_USER,
                    )
                    db.add(contact)
                    db.flush()
                    contacts[key(name)] = contact
                    created_contacts += 1
                exists = db.scalar(select(ProjectContact).where(
                    ProjectContact.project_id == project.id,
                    ProjectContact.contact_id == contact.id,
                    ProjectContact.project_role == "Client",
                ))
                if not exists:
                    db.add(ProjectContact(
                        project=project, contact=contact, project_role="Client",
                        notes=f"Linked from {SOURCE}", created_by=AUDIT_USER, updated_by=AUDIT_USER,
                    ))
                    links += 1

        print(f"Workbook rows: {sum(len(rows) for rows in grouped.values())}")
        print(f"Unique jobsites: {len(grouped)}")
        print(f"Properties to create: {created_properties}")
        print(f"Projects to create: {created_projects}")
        print(f"Existing projects to enrich: {updated_projects}")
        print(f"New name-only contacts: {created_contacts}")
        print(f"Client links to add: {links}")
        if args.commit:
            db.commit()
            print("Committed import.")
        else:
            db.rollback()
            print("Dry run only; no database changes made.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
