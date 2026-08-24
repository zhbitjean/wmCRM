import csv
import io
from pathlib import Path
from openpyxl import load_workbook
from .geography import infer_nyc_borough

EXPECTED_HEADERS = {
    "name", "nickname", "company", "address", "office phone",
    "fax", "cell phone", "notes", "email",
}

def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return None if not text or text in {"-", "—"} else text

def split_name(name):
    parts = (name or "").split()
    if not parts:
        return "", ""
    return parts[0], " ".join(parts[1:])

def map_client_row(row):
    normalized = {str(k).strip().lower(): clean(v) for k, v in row.items() if k is not None}
    name = normalized.get("name")
    if not name:
        return None
    first, last = split_name(name)
    cell = normalized.get("cell phone")
    office = normalized.get("office phone")
    return {
        "display_name": name,
        "first_name": first,
        "last_name": last,
        "nickname": normalized.get("nickname"),
        "company_name": normalized.get("company"),
        "company_address": normalized.get("address"),
        "company_phone": office,
        "company_fax": normalized.get("fax"),
        "address": normalized.get("address"),
        "borough": infer_nyc_borough(normalized.get("address")),
        "fax": normalized.get("fax"),
        "phone": cell or office,
        "alternate_phone": office if cell and office else None,
        "email": normalized.get("email"),
        "role": "Client",
        "notes": normalized.get("notes"),
    }

def parse_current_client_xlsx(content: bytes):
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if "current client list" not in workbook.sheetnames:
            raise ValueError("Workbook does not contain a 'current client list' sheet")
        sheet = workbook["current client list"]
        rows = sheet.iter_rows(values_only=True)
        headers = [clean(value) or "" for value in next(rows, ())]
        found = {header.lower() for header in headers}
        missing = EXPECTED_HEADERS - found
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")
        mapped, skipped = [], []
        for number, values in enumerate(rows, start=2):
            row = dict(zip(headers, values))
            if not any(clean(value) for value in values):
                continue
            item = map_client_row(row)
            if item:
                mapped.append(item)
            else:
                skipped.append({"row": number, "reason": "Name is blank", "nickname": clean(row.get("Nickname"))})
        return mapped, skipped
    finally:
        workbook.close()

def parse_csv(content: bytes):
    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))
